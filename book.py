from openpyxl import Workbook, load_workbook


class Excel:
    def __init__(self) -> None:
        self.wb = Workbook()
        self.wb.save('information.xlsx')

    def __load_book__(self) -> None:
        self.wb = load_workbook('information.xlsx')
        self.ws = self.wb.active

    def __save_book__(self) -> None:
        self.wb.save('information.xlsx')

    def __name_line__(self) -> None:  # функция для создания шапки в excel
        self.__load_book__()
        self.ws["A1"] = "Название"
        self.ws["B1"] = "Ссылка"
        self.ws["C1"] = "Полная цена"
        self.ws["D1"] = "цена за кв.м"
        self.ws["E1"] = "Адресс"
        self.ws["F1"] = "Количество комнат"
        self.ws["G1"] = "Общая площадь"
        self.ws["H1"] = "Площадь кухни"
        self.ws["I1"] = "Жилая площадь"
        self.ws["J1"] = "Этаж"
        self.ws["K1"] = "Балкон/лоджия"
        self.ws["L1"] = "Тип комнат"
        self.ws["M1"] = "Высота потолков"
        self.ws["N1"] = "Санузел"
        self.ws["O1"] = "Окна"
        self.ws["P1"] = "Ремонт"
        self.ws["Q1"] = "Мебель"
        self.ws["R1"] = "Способ продажи"
        self.ws["S1"] = "Отделка"
        self.ws["T1"] = "Вид сделки"
        self.ws["U1"] = "Техника"
        self.__save_book__()
    # функция записи данных в ячейки таблицы, принимает аргументы и список

    def __with_book__(self, iteration: int, name: str, url: str, price: str, sell: str, adres: str, discription: list) -> None:
        iteration += 1
        self.__load_book__()
        self.ws[f"A{iteration}"] = name
        self.ws[f"B{iteration}"] = url
        self.ws[f"C{iteration}"] = price
        self.ws[f"D{iteration}"] = sell
        self.ws[f"E{iteration}"] = adres
        for line in discription:  # цикл для удаление артикулов из строк списка и запись в ячейки
            if "Количество комнат" in line:
                self.ws[f"F{iteration}"] = line.replace(
                    "Количество комнат: ", "")
            elif "Общая площадь" in line:
                self.ws[f"G{iteration}"] = line.replace("Общая площадь: ", "")
            elif "Площадь кухни" in line:
                self.ws[f"H{iteration}"] = line.replace("Площадь кухни: ", "")
            elif "Жилая площадь" in line:
                self.ws[f"I{iteration}"] = line.replace("Жилая площадь: ", "")
            elif "Этаж" in line:
                self.ws[f"J{iteration}"] = line.replace("Этаж: ", "")
            elif "Балкон или лоджия" in line:
                self.ws[f"K{iteration}"] = line.replace(
                    "Балкон или лоджия: ", "")
            elif "Тип комнат" in line:
                self.ws[f"L{iteration}"] = line.replace("Тип комнат: ", "")
            elif "Высота потолков" in line:
                self.ws[f"M{iteration}"] = line.replace(
                    "Высота потолков: ", "")
            elif "Санузел" in line:
                self.ws[f"N{iteration}"] = line.replace("Санузел: ", "")
            elif "Окна" in line:
                self.ws[f"O{iteration}"] = line.replace("Окна: ", "")
            elif "Ремонт" in line:
                self.ws[f"P{iteration}"] = line.replace("Ремонт: ", "")
            elif "Мебель" in line:
                self.ws[f"Q{iteration}"] = line.replace("Мебель: ", "")
            elif "Способ продажи" in line:
                self.ws[f"R{iteration}"] = line.replace("Способ продажи: ", "")
            elif "Отделка" in line:
                self.ws[f"S{iteration}"] = line.replace("Отделка: ", "")
            elif "Вид сделки" in line:
                self.ws[f"T{iteration}"] = line.replace("Вид сделки: ", "")
            elif "Техника" in line:
                self.ws[f"U{iteration}"] = line.replace("Техника: ", "")
            else:
                self.ws[f"V{iteration}"] = line

        self.__save_book__()
